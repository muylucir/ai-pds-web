# backend/pathfinder/parsers/uploads.py
"""업로드 파일 → 에이전트가 file_read로 읽을 텍스트 변환.

xlsx는 VM 안 에이전트가 직접 못 읽으므로(텍스트 도구뿐) 업로드 시점에
마크다운 표로 변환한다. 변환 결과는 룰의 URL 모드와 같은 50,000자 한도로
절단한다(spec §6). 내용은 신뢰하지 않는 입력 — 텍스트로만 저장한다.
"""
from __future__ import annotations
import io
import re

MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_CHARS = 50_000
_TRUNC_MARK = "\n[... 50,000자 초과분 생략]"
ALLOWED = {".md", ".txt", ".csv", ".xlsx", ".pdf"}


def _ext(filename: str) -> str:
    dot = filename.rfind(".")
    return filename[dot:].lower() if dot >= 0 else ""


def _xlsx_to_markdown(data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        rows = [[("" if c is None else str(c)) for c in row]
                for row in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        parts.append(f"## {ws.title}")
        parts.append("| " + " | ".join(rows[0]) + " |")
        parts.append("|" + "---|" * len(rows[0]))
        for row in rows[1:]:
            parts.append("| " + " | ".join(row) + " |")
    return "\n".join(parts)


def _pdf_to_text(data: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(data))
    return "\n\n".join((page.extract_text() or "") for page in reader.pages)


def convert(filename: str, data: bytes) -> tuple[str, bool]:
    ext = _ext(filename)
    if ext not in ALLOWED:
        raise ValueError(f"unsupported extension: {ext or '(none)'}")
    if ext == ".xlsx":
        try:
            content = _xlsx_to_markdown(data)
        except Exception as e:
            # A valid extension with corrupt/non-xlsx bytes (zipfile.BadZipFile,
            # KeyError from openpyxl, etc.) must not surface as a raw 500 — the
            # route's existing ValueError->415 handler is the intended path for
            # any "we can't make sense of this upload" outcome.
            raise ValueError(f"cannot parse {ext} file: corrupted or invalid") from e
    elif ext == ".pdf":
        try:
            content = _pdf_to_text(data)
        except Exception as e:
            raise ValueError(f"cannot parse {ext} file: corrupted or invalid") from e
    else:  # .md .txt .csv — 텍스트 그대로 (lossy 디코드)
        content = data.decode("utf-8", errors="replace")
    if len(content) > MAX_CHARS:
        cut = MAX_CHARS - len(_TRUNC_MARK)
        return content[:cut] + _TRUNC_MARK, True
    return content, False


def safe_name(filename: str, existing: set[str]) -> str:
    """원본 이름을 워크스페이스 안전 슬러그로. 한글 유지, 경로·특수문자 제거,
    확장자는 항상 .md(변환 결과물이므로). 충돌 시 -2, -3… 접미사."""
    stem = filename
    dot = stem.rfind(".")
    if dot > 0:
        stem = stem[:dot]
    stem = re.sub(r"[^\w가-힣-]+", "-", stem).strip("-") or "upload"
    candidate = f"{stem}.md"
    n = 2
    while candidate in existing:
        candidate = f"{stem}-{n}.md"
        n += 1
    return candidate
